from flask import Flask, render_template, request, send_file
import pandas as pd
import random
import xlsxwriter
import openpyxl
import os

app = Flask(__name__)

# Constants
TOTAL_HRS = 7
DAYS = 5
MAX_SIZE = 120
GAP = 17
classes = []

def populate_teacher(s):
    length = len(s)
    class_ind = {}
    for k in range(length):
        if str(s.iat[k, 1]).lower() != 'nan':
            class_ind[s.iat[k, 0]] = s.iat[k, 1].split(',')
        else:
            classes.append(str(s.iat[k, 0]))
    return class_ind

def populate(s):
    length = len(s)
    list_1 = []
    class_ind = []
    for k in range(length):
        if str(s.iat[k, 1]).lower() != 'nan':
            class_ind.append(list((s.iat[k, 0], s.iat[k, 1])))
        else:
            if class_ind:
                list_1.append(class_ind)
            classes.append(str(s.iat[k, 0]))
            class_ind = []
    list_1.append(class_ind)
    return list_1

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/page2.html')
def page2():
    return render_template('page2.html')

@app.route('/download', methods=['POST'])
def view():
    files = [request.files[f'file{i}'] for i in range(1, 4)]
   
    s2 = pd.read_excel(files[0], skiprows=2)
    s1 = pd.read_excel(files[1])
    s3 = pd.read_excel(files[2])
   
    path = files[0]
    wb_obj = openpyxl.load_workbook(path)  
    sheet_obj = wb_obj.active
    cell_obj1 = str(sheet_obj.cell(row=1, column=1).value)
    cell_obj2 = str(sheet_obj.cell(row=2, column=1).value)
   
    # Save the output file in the /tmp directory
    output_path = os.path.join('/tmp', 'final.xlsx')
    wb = xlsxwriter.Workbook(output_path)
    ws = wb.add_worksheet("TimeTable")
    ws2 = wb.add_worksheet("TeacherSlot")
    f2 = wb.add_format({'bold': True, 'bg_color': '#b2b2b2'})
    f3 = wb.add_format({'bg_color': '#808080'})
    f4 = wb.add_format({'bold': True, 'bg_color': '#808080'})
    f5 = wb.add_format({'bg_color': '#b2b2b2'})
    f6 = wb.add_format({'bold': True, 'bg_color': '#999999'})
    f7 = wb.add_format({'bold': True})
    working_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    teachers = s1['faculty'].dropna().unique().tolist()
    tt = []
    for teacher in teachers:
        a = teacher.split(',')
        if len(a) == 1:
            tt.append(a[0])
        else:
            for i in a:
                tt.append(i)
    teachers = list(set(tt))

    teacher_course = populate_teacher(s1)
    teacher_len = len(teachers)

    t_len = len(s2)
    course_hour = populate(s3)
   
    timeslot = [[0]*MAX_SIZE for _ in range(t_len)]
    teacherslot = [[0]*MAX_SIZE for _ in range(teacher_len)]

    for i in range(t_len):
        index = 0
        for k in range(DAYS):
            for j in range(TOTAL_HRS):
                timeslot[i][index] = str(s2.iat[i, (k*TOTAL_HRS)+j])
                if timeslot[i][index] in teacher_course:
                    fac = teacher_course[timeslot[i][index]]
                    t_index = []
                    for teacher in fac:
                        t_index.append(teachers.index(teacher))
                    for tindex in t_index:
                        teacherslot[tindex][index] = timeslot[i][index]
                index += 1
            index += GAP

    for k in range(t_len):
        c_h = course_hour[k]
        for i in range(len(c_h)):
            course = c_h[i][0]
            hour = c_h[i][1]
            fac = teacher_course[course]
            t_index = []
            for teacher in fac:
                t_index.append(teachers.index(teacher))
            rem_hr = hour
            slots = []
            while int(rem_hr) > 0:
                for j in range(MAX_SIZE):
                    if str(timeslot[k][j]).lower() == "nan":
                        begin = j
                        break
                for j in range((MAX_SIZE-1), -1, -1):
                    if str(timeslot[k][j]).lower() == "nan":
                        end = j
                        break
                if rem_hr != 1:
                    interval = (end-begin+1)/(rem_hr-1)
                pos = begin
                for j in range(int(rem_hr)):
                    slots.append(pos)
                    pos = pos + interval
                for slot in slots:
                    flag = 0
                    flag2 = 0
                    if str(timeslot[k][int(slot)]).lower() == "nan":
                        for tindex in t_index:
                            if teacherslot[tindex][int(slot)] != 0 or teacherslot[tindex][(int(slot)-1) % MAX_SIZE] != 0 or teacherslot[tindex][(int(slot)+1) % MAX_SIZE] != 0:
                                flag = 1
                        if flag == 0:
                            timeslot[k][int(slot)] = course
                            for tindex in t_index:
                                teacherslot[tindex][int(slot)] = course
                            flag2 = 1
                    elif flag2 == 0:
                        left = int(slot) - 1
                        right = int(slot) + 1
                        while left > 0 or right < MAX_SIZE:
                            if ((left > 0 and str(timeslot[k][left]).lower() == "nan")):
                                f = 0
                                for tindex in t_index:
                                    if teacherslot[tindex][left] != 0 or teacherslot[tindex][(left-1) % MAX_SIZE] != 0 or teacherslot[tindex][(left+1) % MAX_SIZE]:
                                        f = 1
                                if f == 0:
                                    timeslot[k][left] = course
                                    for tindex in t_index:
                                        teacherslot[tindex][left] = course
                                    break
                            if (right < MAX_SIZE and str(timeslot[k][right]).lower() == "nan"):
                                f = 0
                                for tindex in t_index:
                                    if teacherslot[tindex][right] != 0 or teacherslot[tindex][(right-1) % MAX_SIZE] != 0 or teacherslot[tindex][(right+1) % MAX_SIZE]:
                                        f = 1
                                if f == 0:
                                    timeslot[k][right] = course
                                    for tindex in t_index:
                                        teacherslot[tindex][right] = course
                                    break
                            left = left - 1
                            right = right + 1
                        if left < 0 and right >= MAX_SIZE:
                            print("ERROR: ALLOCATION COULD NOT BE DONE")
                            break
                    rem_hr -= 1
                else:
                    continue
                break
           
    k = 0
    timetable = []
    counter = 3
   
    merge_format = wb.add_format({
        "bold": 1,
        "align": "center",
        "valign": "vcenter"
    })
    ws.merge_range("A1:U1", cell_obj1, merge_format)
    ws.merge_range("A2:U2", cell_obj2, merge_format)
   
    while k < t_len:
        index = 0
        temp = [[0]*TOTAL_HRS for _ in range(DAYS)]
        timetable.append(['', '', '', classes[k], '', '', ''])
        for i in range(DAYS):
            for j in range(TOTAL_HRS):
                if str(timeslot[k][index]).lower() == "nan":
                    timeslot[k][index] = "REMEDIAL"
                temp[i][j] = timeslot[k][index]
                index += 1
            timetable.append(temp[i])
            index += GAP
        ws.write(counter, 4, classes[k])
        if k % 2 == 0:
            ws.write_row(counter+1, 0, ['', '1st', '2nd', '3rd', 'Lunch', '4th', '5th', '6th'], f6)
        else:
            ws.write_row(counter+1, 0, ['', '1st', '2nd', '3rd', 'Lunch', '4th', '5th', '6th'], f7)
        for i in range(DAYS):
            if i % 2 == 0:
                ws.write(counter+2+i, 0, working_days[i], f4)
                ws.write_row(counter+2+i, 1, temp[i], f5)
            else:
                ws.write(counter+2+i, 0, working_days[i], f3)
                ws.write_row(counter+2+i, 1, temp[i])
        counter += 7
        k += 1
   
    teacherslot_table = []
    for i in range(teacher_len):
        index = 0
        temp = [[0]*TOTAL_HRS for _ in range(DAYS)]
        teacherslot_table.append([teachers[i]])
        for day in range(DAYS):
            for hr in range(TOTAL_HRS):
                if str(teacherslot[i][index]).lower() == "nan":
                    teacherslot[i][index] = ""
                temp[day][hr] = teacherslot[i][index]
                index += 1
            teacherslot_table.append(temp[day])
            index += GAP
           
    counter = 2
    for i in range(teacher_len):
        index = 0
        temp = [[0]*TOTAL_HRS for _ in range(DAYS)]
        teacherslot_table.append([teachers[i]])
        ws2.write(counter, 0, teachers[i], f2)
        for day in range(DAYS):
            if day % 2 == 0:
                ws2.write(counter+1, 0, working_days[day], f4)
            else:
                ws2.write(counter+1, 0, working_days[day], f3)
            for hr in range(TOTAL_HRS):
                temp[day][hr] = teacherslot[i][index]
                index += 1
            ws2.write_row(counter+1, 1, temp[day], f5)
            index += GAP
            counter += 1
        counter += 1
       
    wb.close()
   
    # Send the file for download
    return send_file(output_path, as_attachment=True, download_name="final.xlsx")

# if __name__ == "__main__":
#     app.run(debug=True)
